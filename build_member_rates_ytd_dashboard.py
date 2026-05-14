#!/usr/bin/env python3
"""
Build the Member Rates YTD dashboard HTML for My Place Hotels.

Reads:
  - ROLLUP_XLSX  (Channel x MemberType rollup totals, sheet "Export")
  - BYHOTEL_XLSX (Property x Channel x MemberType breakout, sheet "Export")
  - MASTER_XLSX  (Hotel master DB; col A=Hotel ID, B=Hotel Name,
                  J=Member Rates pilot flag ("X"/"x"=pilot, blank=non-pilot))

Computes three cohort aggregations:
  - Total Portfolio  : all hotels in the by-hotel file (matches rollup totals)
  - Pilot Cohort     : just the 32 hotels flagged X in the master DB
  - Non-Pilot Cohort : the remaining 49 hotels

Rewrites the dashboard HTML with:
  - Cohort toggle (Total / Pilot Only / Non-Pilot Only / Side-by-Side)
  - Pilot vs New badges in the hotel table
  - A "Cohort" column in the hotel table
  - Updated title / subtitle / table title

To roll forward month-end:
  1. Drop in fresh `2026-ytd-memberrates-MM-DD-YY.xlsx` +
     `2026-ytd-memberrates-byhotel-MM-DD-YY.xlsx`
  2. Update the three filename constants below
  3. `python3 build_member_rates_ytd_dashboard.py`
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:
    sys.stderr.write("openpyxl is required: install with `pip3 install openpyxl`\n")
    raise


# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────
ROLLUP_FILE  = "2026-ytd-memberrates-4-30-26.xlsx"
BYHOTEL_FILE = "2026-ytd-memberrates-byhotel-4-30-26.xlsx"
MASTER_FILE  = "My Place Hotels Database - Master.xlsx"
OUTPUT_HTML  = "2026-ytd-member-rates-dashboard-branded.html"

# Period label used in subtitle / footer text
PERIOD_LABEL = "Jan 1 – Apr 30, 2026"
DATA_AS_OF   = "Apr 30, 2026"

# Resolve BASE: prefer macOS path, else fall back to the cowork workspace mount,
# else the folder this script lives in.
_CANDIDATE_BASES = [
    Path("/Users/jeffthomas/Documents/Claude/Projects/MPH Performance Dashboards"),
    Path("/sessions/admiring-youthful-faraday/mnt/MPH Performance Dashboards"),
    Path(__file__).resolve().parent,
]
BASE = next((p for p in _CANDIDATE_BASES if p.exists()), _CANDIDATE_BASES[-1])

ROLLUP_PATH  = BASE / ROLLUP_FILE
BYHOTEL_PATH = BASE / BYHOTEL_FILE
MASTER_PATH  = BASE / MASTER_FILE
OUTPUT_PATH  = BASE / OUTPUT_HTML


# ─────────────────────────────────────────────────────────────────────────────
# Name normalization & pilot-flag map
# ─────────────────────────────────────────────────────────────────────────────
def normalize(s) -> str:
    if s is None:
        return ""
    s = str(s).lower()
    # Common typos & variants between master DB and by-hotel exports
    s = s.replace("plainfeild", "plainfield")
    s = s.replace("aiport", "airport")
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def strip_state_suffix(k: str) -> str:
    """Drop trailing 2-letter state from a normalized key, if present."""
    if len(k) > 4 and re.match(r"^[a-z]{2}$", k[-2:]):
        return k[:-2]
    return k


# Manual aliases: by-hotel normalized key -> master normalized key
MANUAL_ALIASES = {
    "bentonvillerogersar": "bentonvillear",
}


def load_master_pilot_map() -> dict:
    """Return {normalized_property_key: {'name': display, 'pilot': bool,
                                           'mgmt': str, 'hotel_id': id}}.

    Column layout (May 2026 master DB):
       A=Hotel ID (idx 0), B=Hotel Name (1), H=Management Company (7),
       I=Member Rates Pilot flag (8).
    """
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True)
    ws = wb["Master Database"]
    master = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None or row[1] is None:
            continue
        name = str(row[1]).strip()
        if not name:
            continue
        pilot = bool(row[8] and str(row[8]).strip().lower() == "x")  # Col I (0-indexed = 8) in May 2026 DB layout
        mgmt_raw = row[7] if len(row) > 7 else None  # Col H (0-indexed = 7)
        mgmt = (str(mgmt_raw).strip() if mgmt_raw is not None and str(mgmt_raw).strip() else "Unknown")
        short = re.sub(r"^My Place Hotel\s*-\s*", "", name).strip()
        key = normalize(short)
        master[key] = {"name": name, "short": short, "pilot": pilot,
                       "mgmt": mgmt, "hotel_id": row[0]}
    return master


def resolve_pilot(prop_name: str, master: dict, master_stripped: dict, warnings: list) -> bool:
    k = normalize(prop_name)
    if k in MANUAL_ALIASES:
        k = MANUAL_ALIASES[k]
    if k in master:
        return master[k]["pilot"]
    k2 = strip_state_suffix(k)
    if k2 in master:
        return master[k2]["pilot"]
    if k2 in master_stripped:
        return master_stripped[k2]["pilot"]
    if k in master_stripped:
        return master_stripped[k]["pilot"]
    warnings.append(f"Property {prop_name!r} did not match master DB; defaulting to Non-Pilot")
    return False


def resolve_mgmt(prop_name: str, master: dict, master_stripped: dict, warnings: list) -> str:
    """Look up Management Company for a by-hotel property name. Falls back to 'Unknown'."""
    k = normalize(prop_name)
    if k in MANUAL_ALIASES:
        k = MANUAL_ALIASES[k]
    if k in master:
        return master[k]["mgmt"]
    k2 = strip_state_suffix(k)
    if k2 in master:
        return master[k2]["mgmt"]
    if k2 in master_stripped:
        return master_stripped[k2]["mgmt"]
    if k in master_stripped:
        return master_stripped[k]["mgmt"]
    warnings.append(f"Property {prop_name!r} did not match master DB; defaulting Management Company to Unknown")
    return "Unknown"


# ─────────────────────────────────────────────────────────────────────────────
# Read rollup file
# ─────────────────────────────────────────────────────────────────────────────
def read_rollup() -> dict:
    """Return {('PMS'|'WEB'|'Total', 'Total'|'Current Member'|'New Member'|'Non-Member'): row_dict}.

    The rollup sheet:
       row 1 = header
       Channel column carries 'PMS', 'WEB', 'Total' on Total-rows; MemberType
       column carries 'Total'/'Current Member'/'New Member'/'Non-Member'/None.
    """
    wb = openpyxl.load_workbook(ROLLUP_PATH, data_only=True)
    ws = wb["Export"]
    out = {}
    current_channel = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is not None and str(row[0]).startswith("Applied filters"):
            break
        if row[0] is not None:
            current_channel = str(row[0]).strip()
        chan = current_channel
        mt = row[1]
        if mt is None:
            # untyped sub-row (Member-Rates rate with no member-type metadata)
            mt_key = "Unknown"
        else:
            mt_key = str(mt).strip()
        if chan is None:
            continue
        out[(chan, mt_key)] = {
            "guests": row[2] or 0,
            "gross_res": row[3] or 0,
            "net_res": row[4] or 0,
            "gross_rn": row[5] or 0,
            "net_rn": row[6] or 0,
            "gross_rev": float(row[7] or 0),
            "net_rev": float(row[8] or 0),
            "gross_adr": float(row[9] or 0),
            "net_adr": float(row[10] or 0),
            "net_rpb": float(row[11] or 0),
            "net_los": float(row[12] or 0),
            "cancel_rate": float(row[13] or 0),
        }
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Read by-hotel file
# ─────────────────────────────────────────────────────────────────────────────
def read_byhotel() -> list:
    """Return a flat list of records: one row per (property, channel, member_type)."""
    wb = openpyxl.load_workbook(BYHOTEL_PATH, data_only=True)
    ws = wb["Export"]
    records = []
    cur_prop = None
    cur_chan = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is not None and str(row[0]).startswith("Applied filters"):
            break
        if row[0] is not None:
            cur_prop = str(row[0]).strip()
            cur_chan = None
        if cur_prop is None:
            continue
        chan = row[1]
        mt = row[2]
        if chan is not None:
            cur_chan = str(chan).strip()
        if cur_chan in (None, "Total"):
            continue
        # member-type row under PMS/WEB
        if mt is None:
            mt_label = "Unknown"
        else:
            mt_label = str(mt).strip()
        records.append({
            "prop": cur_prop,
            "chan": cur_chan,
            "mt": mt_label,
            "guests": row[3] or 0,
            "gross_res": row[4] or 0,
            "net_res": row[5] or 0,
            "gross_rn": row[6] or 0,
            "net_rn": row[7] or 0,
            "gross_rev": float(row[8] or 0),
            "net_rev": float(row[9] or 0),
            "net_adr_local": float(row[11] or 0),
            "net_los_local": float(row[13] or 0),
        })
    return records


# ─────────────────────────────────────────────────────────────────────────────
# Aggregation
# ─────────────────────────────────────────────────────────────────────────────
def aggregate(records, properties_in_scope=None) -> dict:
    """Aggregate by (channel, member_type) across the given properties.

    Returns a dict mirroring the structure consumed by the dashboard JS:
       {
         'pms': {'total': {...}, 'current': {...}, 'new_mem': {...}, 'non_mem': {...}},
         'web': {'total': {...}, 'current': {...}, 'new_mem': {...}, 'non_mem': {...}},
         'grand': {...}
       }
    """
    # Initialise accumulators
    def empty():
        return {
            "guests": 0, "gross_res": 0, "net_res": 0,
            "gross_rn": 0, "net_rn": 0,
            "gross_rev": 0.0, "net_rev": 0.0,
        }

    buckets = {
        ("PMS", "Current Member"): empty(),
        ("PMS", "New Member"):    empty(),
        ("PMS", "Non-Member"):    empty(),
        ("PMS", "Unknown"):       empty(),
        ("WEB", "Current Member"): empty(),
        ("WEB", "New Member"):    empty(),
        ("WEB", "Non-Member"):    empty(),
        ("WEB", "Unknown"):       empty(),
    }
    for r in records:
        if properties_in_scope is not None and r["prop"] not in properties_in_scope:
            continue
        key = (r["chan"], r["mt"])
        if key not in buckets:
            continue
        b = buckets[key]
        for f in ("guests", "gross_res", "net_res", "gross_rn", "net_rn", "gross_rev", "net_rev"):
            b[f] += r[f]

    def fold_channel(chan):
        cm = buckets[(chan, "Current Member")]
        nm = buckets[(chan, "New Member")]
        nom = buckets[(chan, "Non-Member")]
        unk = buckets[(chan, "Unknown")]
        tot = empty()
        for src in (cm, nm, nom, unk):
            for f in tot:
                tot[f] += src[f]

        def derived(b):
            b = dict(b)
            b["net_adr"] = (b["net_rev"] / b["net_rn"]) if b["net_rn"] else 0.0
            b["gross_adr"] = (b["gross_rev"] / b["gross_rn"]) if b["gross_rn"] else 0.0
            b["net_rpb"] = (b["net_rev"] / b["net_res"]) if b["net_res"] else 0.0
            b["net_los"] = (b["net_rn"] / b["net_res"]) if b["net_res"] else 0.0
            b["cancel_rate"] = (1 - (b["net_res"] / b["gross_res"])) if b["gross_res"] else 0.0
            return b
        return {
            "total":   derived(tot),
            "current": derived(cm),
            "new_mem": derived(nm),
            "non_mem": derived(nom),
        }

    pms = fold_channel("PMS")
    web = fold_channel("WEB")

    grand = {
        "guests":     pms["total"]["guests"]     + web["total"]["guests"],
        "gross_res":  pms["total"]["gross_res"]  + web["total"]["gross_res"],
        "net_res":    pms["total"]["net_res"]    + web["total"]["net_res"],
        "gross_rn":   pms["total"]["gross_rn"]   + web["total"]["gross_rn"],
        "net_rn":     pms["total"]["net_rn"]     + web["total"]["net_rn"],
        "gross_rev":  pms["total"]["gross_rev"]  + web["total"]["gross_rev"],
        "net_rev":    pms["total"]["net_rev"]    + web["total"]["net_rev"],
    }
    grand["net_adr"]     = (grand["net_rev"] / grand["net_rn"]) if grand["net_rn"] else 0.0
    grand["gross_adr"]   = (grand["gross_rev"] / grand["gross_rn"]) if grand["gross_rn"] else 0.0
    grand["net_rpb"]     = (grand["net_rev"] / grand["net_res"]) if grand["net_res"] else 0.0
    grand["net_los"]     = (grand["net_rn"] / grand["net_res"]) if grand["net_res"] else 0.0
    grand["cancel_rate"] = (1 - (grand["net_res"] / grand["gross_res"])) if grand["gross_res"] else 0.0

    return {"pms": pms, "web": web, "grand": grand}


def build_hotel_rows(records, pilot_map_by_prop, mgmt_map_by_prop) -> list:
    """Build the per-hotel rows shown in the table.

    Each row carries:
      - Display fields used by the table (revenue/res totals by channel + member type)
      - `cells`: full raw breakdown by (channel, member_type) of all summable
        fields. JS sums `cells` across the filtered hotel set to reconstruct
        the cohort aggregation structure on the fly. This lets cohort and
        management-company filters compose independently.
    """
    # Raw fields summed per (channel, member_type) cell — derived metrics
    # like ADR/RPB/LOS are computed in JS after summation.
    RAW_FIELDS = ("guests", "gross_res", "net_res", "gross_rn", "net_rn",
                  "gross_rev", "net_rev")
    MT_KEYS = {"Current Member": "current", "New Member": "new_mem",
               "Non-Member": "non_mem", "Unknown": "unknown"}

    by_prop = {}
    for r in records:
        p = r["prop"]
        slot = by_prop.setdefault(p, {
            "guests": 0,
            "pms_cm_rev": 0.0, "pms_cm_res": 0,
            "pms_nm_rev": 0.0, "pms_nm_res": 0,
            "web_cm_rev": 0.0, "web_cm_res": 0,
            "web_nm_rev": 0.0, "web_nm_res": 0,
            "net_rev": 0.0, "net_res": 0, "net_rn": 0,
            "pms_total_rev": 0.0, "pms_total_res": 0,
            "web_total_rev": 0.0, "web_total_res": 0,
            # Full cell grid: cells[chan][mt_key][raw_field]
            "cells": {
                "pms": {k: {f: 0 for f in RAW_FIELDS} for k in MT_KEYS.values()},
                "web": {k: {f: 0 for f in RAW_FIELDS} for k in MT_KEYS.values()},
            },
        })
        slot["guests"] += r["guests"]
        slot["net_rev"] += r["net_rev"]
        slot["net_res"] += r["net_res"]
        slot["net_rn"]  += r["net_rn"]

        chan_key = "pms" if r["chan"] == "PMS" else ("web" if r["chan"] == "WEB" else None)
        # Drop the per-channel "Total" sub-row (it's the sum of the
        # member-type rows that follow and would double-count). Only the
        # member-type detail rows feed the cells grid.
        if chan_key is not None and r["mt"] in MT_KEYS:
            mt_key = MT_KEYS[r["mt"]]
            cell = slot["cells"][chan_key][mt_key]
            for f in RAW_FIELDS:
                cell[f] += r[f]

        if r["chan"] == "PMS":
            slot["pms_total_rev"] += r["net_rev"]
            slot["pms_total_res"] += r["net_res"]
            if r["mt"] == "Current Member":
                slot["pms_cm_rev"] += r["net_rev"]; slot["pms_cm_res"] += r["net_res"]
            elif r["mt"] == "New Member":
                slot["pms_nm_rev"] += r["net_rev"]; slot["pms_nm_res"] += r["net_res"]
        elif r["chan"] == "WEB":
            slot["web_total_rev"] += r["net_rev"]
            slot["web_total_res"] += r["net_res"]
            if r["mt"] == "Current Member":
                slot["web_cm_rev"] += r["net_rev"]; slot["web_cm_res"] += r["net_res"]
            elif r["mt"] == "New Member":
                slot["web_nm_rev"] += r["net_rev"]; slot["web_nm_res"] += r["net_res"]

    def _round_cells(cells):
        out = {"pms": {}, "web": {}}
        for chan in ("pms", "web"):
            for mt, b in cells[chan].items():
                out[chan][mt] = {f: (round(v, 4) if isinstance(v, float) else v) for f, v in b.items()}
        return out

    rows = []
    for prop, s in by_prop.items():
        pilot = pilot_map_by_prop.get(prop, False)
        mgmt = mgmt_map_by_prop.get(prop, "Unknown")
        display_name = _prettify_prop_name(prop)
        net_adr = (s["net_rev"] / s["net_rn"]) if s["net_rn"] else 0.0
        rows.append({
            "name": display_name,
            "raw_name": prop,
            "cohort": "Pilot" if pilot else "New",
            "pilot": pilot,
            "mgmt_co": mgmt,
            "guests": s["guests"],
            "pms_cm_rev": round(s["pms_cm_rev"], 2),
            "pms_cm_res": s["pms_cm_res"],
            "pms_nm_rev": round(s["pms_nm_rev"], 2),
            "pms_nm_res": s["pms_nm_res"],
            "web_cm_rev": round(s["web_cm_rev"], 2),
            "web_cm_res": s["web_cm_res"],
            "web_nm_rev": round(s["web_nm_rev"], 2),
            "web_nm_res": s["web_nm_res"],
            "net_adr": round(net_adr, 2),
            "cells": _round_cells(s["cells"]),
        })
    rows.sort(key=lambda r: r["name"].lower())
    return rows


def _prettify_prop_name(p: str) -> str:
    """ALL CAPS by-hotel property -> Title Case for display."""
    # Don't title-case 2-letter state codes after commas.
    # Simple approach: split on commas; title-case each piece, except 2-letter state at end.
    parts = [pt.strip() for pt in p.split(",")]
    out_parts = []
    for idx, pt in enumerate(parts):
        if idx == len(parts) - 1 and len(pt) == 2 and pt.isalpha():
            out_parts.append(pt.upper())
        else:
            # title-case but preserve slashes / dashes
            tokens = re.split(r"([/\-\s])", pt)
            tokens = [t.capitalize() if t.isalpha() else t for t in tokens]
            out_parts.append("".join(tokens).strip())
    return ", ".join(out_parts)


# ─────────────────────────────────────────────────────────────────────────────
# HTML generation
# ─────────────────────────────────────────────────────────────────────────────
LOGO_DATA_URI = ""  # filled in at runtime from logo_uri.txt if present


def _load_logo() -> str:
    p = BASE / "logo_uri.txt"
    if p.exists():
        return p.read_text().strip()
    return ""


def build_html(cohorts: dict, hotel_rows: list, cohort_counts: dict,
               mgmt_options: list, warnings: list) -> str:
    """cohorts = {'total': agg, 'pilot': agg, 'nonpilot': agg}; hotel_rows = list;
       mgmt_options = [{'name': str, 'count': int}, ...] (already sorted)."""
    logo_uri = _load_logo()
    js_data = {
        "cohorts": cohorts,
        "hotels": hotel_rows,
        "counts": cohort_counts,
        "mgmtOptions": mgmt_options,
        "period": PERIOD_LABEL,
        "asOf": DATA_AS_OF,
    }
    js_data_json = json.dumps(js_data, separators=(",", ":"))

    html = HTML_TEMPLATE \
        .replace("__DATA_JSON__", js_data_json) \
        .replace("__LOGO_URI__", logo_uri) \
        .replace("__PERIOD_LABEL__", PERIOD_LABEL) \
        .replace("__DATA_AS_OF__", DATA_AS_OF) \
        .replace("__TOTAL_COUNT__", str(cohort_counts["total"])) \
        .replace("__PILOT_COUNT__", str(cohort_counts["pilot"])) \
        .replace("__NONPILOT_COUNT__", str(cohort_counts["nonpilot"]))
    return html


def build_mgmt_options(hotel_rows: list) -> list:
    """Return list of {name, count} sorted by count desc, with 'All Companies' pinned first."""
    from collections import Counter
    c = Counter(h["mgmt_co"] for h in hotel_rows)
    total = sum(c.values())
    # Sort: count desc, then name asc; pin 'All Companies' first.
    operators = sorted(c.items(), key=lambda kv: (-kv[1], kv[0]))
    out = [{"name": "All Companies", "count": total}]
    for name, count in operators:
        out.append({"name": name, "count": count})
    return out


# The HTML template — single big string. All chart/render functions are
# rewritten to take an `agg` argument; the cohort toggle re-renders everything.
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>2026 YTD Member Rates — Executive Dashboard</title>
<style>
  :root {
    --bg:           #EEF1F6;
    --card:         #FFFFFF;
    --card-hover:   #F4F7FB;
    --border:       #D0D8E8;
    --text:         #1C355E;
    --muted:        #5A6B84;
    --accent:       #1C355E;   /* PMS — Navy */
    --accent2:      #2E8B57;   /* WEB — Forest Green */
    --accent3:      #E11B22;   /* Alert Red */
    --accent4:      #FF5F00;   /* Orange */
    --positive:     #2E8B57;
    --negative:     #CC3333;
  }
  * { margin:0; padding:0; box-sizing:border-box; }
  body { font-family: Arial, 'Helvetica Neue', Helvetica, sans-serif;
         background: var(--bg); color: var(--text); min-height:100vh; }

  /* ── Top Brand Bar ── */
  .brand-bar { background: #FF5F00; height: 5px; }

  /* ── Header ── */
  .header { padding: 24px 32px 18px; background: #1C355E;
            border-bottom: 3px solid #FF5F00; }
  .header h1 { font-size: 22px; font-weight: 700; letter-spacing: -0.3px;
               color: #FFFFFF; }
  .header .subtitle { color: rgba(255,255,255,0.72); font-size: 13px; margin-top: 4px; }
  .pilot-badge { display: inline-block;
                 background: linear-gradient(135deg, #FF5F00, #cc4800);
                 color: white; font-size: 11px; font-weight: 700;
                 padding: 3px 10px; border-radius: 12px;
                 margin-left: 12px; vertical-align: middle; }

  /* ── Filter Bar (sticky) ── */
  .cohort-toggle-wrap { position: sticky; top: 0; z-index: 30;
                        background: rgba(238,241,246,0.96);
                        backdrop-filter: blur(6px);
                        border-bottom: 1px solid var(--border);
                        padding: 10px 32px; }
  .filter-bar { max-width: 1440px; margin: 0 auto;
                display: flex; flex-wrap: wrap; gap: 10px 22px;
                align-items: center; }
  .filter-group { display: flex; flex-wrap: wrap; gap: 8px;
                  align-items: center; }
  .filter-divider { width: 1px; align-self: stretch;
                    background: var(--border); margin: 2px 4px; }
  .cohort-toggle { display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
  .filter-bar .lbl, .cohort-toggle .lbl {
                        font-size: 11px; text-transform: uppercase;
                        letter-spacing: 0.8px; color: var(--muted);
                        font-weight: 700; margin-right: 4px; }
  .cohort-btn { background: #E4EAF3; color: #1C355E; font-size: 12px;
                font-weight: 700; padding: 7px 16px; border-radius: 20px;
                border: 1px solid var(--border); cursor: pointer;
                font-family: Arial, sans-serif; transition: all 0.18s;
                letter-spacing: 0.2px; }
  .cohort-btn:hover { background: #F4F7FB; }
  .cohort-btn:focus-visible { outline: 2px solid #FF5F00;
                              outline-offset: 2px; }
  .cohort-btn.active { background: #1C355E; color: #FFFFFF;
                       border-color: #1C355E; }
  .cohort-btn .ct-count { opacity: 0.7; font-weight: 600;
                          margin-left: 4px; font-size: 11px; }

  /* Management Company dropdown */
  .mgmt-select-wrap { position: relative; display: inline-flex;
                       align-items: center; min-width: 260px; }
  .mgmt-select { appearance: none; -webkit-appearance: none; -moz-appearance: none;
                  background: #FFFFFF; color: #1C355E;
                  font-family: Arial, sans-serif; font-size: 12px;
                  font-weight: 700; letter-spacing: 0.2px;
                  padding: 7px 32px 7px 14px; border-radius: 20px;
                  border: 1px solid var(--border); cursor: pointer;
                  width: 100%; transition: all 0.18s;
                  background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'><path fill='%231C355E' d='M0 0l5 6 5-6z'/></svg>");
                  background-repeat: no-repeat;
                  background-position: right 12px center; }
  .mgmt-select:hover { background-color: #F4F7FB;
                        border-color: #1C355E; }
  .mgmt-select:focus-visible { outline: 2px solid #FF5F00;
                                outline-offset: 2px;
                                border-color: #1C355E; }
  .mgmt-select.filtered { background-color: #1C355E; color: #FFFFFF;
                          border-color: #1C355E;
                          background-image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'><path fill='%23FFFFFF' d='M0 0l5 6 5-6z'/></svg>"); }
  .mgmt-select option { color: #1C355E; background: #FFFFFF;
                        font-weight: 600; }

  /* Empty-state placeholder when filter intersection is empty */
  .empty-state { background: var(--card); border: 1px dashed var(--border);
                 border-radius: 8px; padding: 38px 24px; text-align: center;
                 color: var(--muted); font-size: 13px; margin-bottom: 24px;
                 box-shadow: 0 1px 4px rgba(28,53,94,0.04); }
  .empty-state .es-title { color: #1C355E; font-size: 15px;
                            font-weight: 700; margin-bottom: 8px;
                            letter-spacing: -0.2px; }
  .empty-state .es-sub { font-size: 12px; line-height: 1.5; }

  /* ── Dashboard Layout ── */
  .dashboard { padding: 24px 32px; max-width: 1440px; margin: 0 auto; }

  /* ── KPI Cards ── */
  .kpi-row { display: grid; grid-template-columns: repeat(6, 1fr);
             gap: 14px; margin-bottom: 24px; }
  .kpi-card { background: var(--card); border: 1px solid var(--border);
              border-top: 3px solid #1C355E;
              border-radius: 8px; padding: 18px 16px;
              transition: box-shadow 0.2s;
              box-shadow: 0 1px 4px rgba(28,53,94,0.07); }
  .kpi-card:hover { box-shadow: 0 4px 14px rgba(28,53,94,0.12);
                    background: var(--card-hover); }
  .kpi-label { font-size: 10px; text-transform: uppercase; letter-spacing: 1px;
               color: var(--muted); font-weight: 700; margin-bottom: 8px; }
  .kpi-value { font-size: 26px; font-weight: 700; letter-spacing: -0.5px;
               color: #1C355E; }
  .kpi-sub { font-size: 11px; color: var(--muted); margin-top: 6px; }
  .kpi-sub span { font-weight: 600; }
  .kpi-sub .pms { color: #1C355E; }
  .kpi-sub .web { color: #2E8B57; }

  /* ── Side-by-side compact KPI strips ── */
  .sxs-wrap { display: grid; grid-template-columns: 1fr 1fr; gap: 16px;
              margin-bottom: 22px; }
  .sxs-col { background: var(--card); border: 1px solid var(--border);
             border-radius: 8px; padding: 16px;
             box-shadow: 0 1px 4px rgba(28,53,94,0.06); }
  .sxs-col.pilot { border-top: 4px solid #1C355E; }
  .sxs-col.newcohort { border-top: 4px solid #FF5F00; }
  .sxs-col h4 { font-size: 12px; text-transform: uppercase; letter-spacing: 0.8px;
                margin-bottom: 12px; color: #1C355E;
                display: flex; align-items: center; gap: 8px; }
  .sxs-col h4 .tag { font-size: 10px; padding: 2px 8px; border-radius: 8px;
                     color:#fff; font-weight: 700; letter-spacing: 0.2px; }
  .sxs-col.pilot h4 .tag { background:#1C355E; }
  .sxs-col.newcohort h4 .tag { background:#FF5F00; }
  .sxs-kpi-grid { display: grid; grid-template-columns: 1fr 1fr 1fr;
                  gap: 8px; }
  .sxs-kpi { padding: 10px; background: #F5F8FC; border-radius: 6px; }
  .sxs-kpi .lbl { font-size: 9px; text-transform: uppercase; color: var(--muted);
                  letter-spacing: 0.5px; font-weight: 700; }
  .sxs-kpi .val { font-size: 18px; font-weight: 700; color: #1C355E;
                  margin-top: 4px; }
  .sxs-kpi .sub { font-size: 10px; color: var(--muted); margin-top: 2px; }

  /* ── Section Titles ── */
  .section-title { font-size: 14px; font-weight: 700; margin-bottom: 14px;
                   display: flex; align-items: center; gap: 8px;
                   color: #1C355E; }
  .section-title .dot { width: 8px; height: 8px; border-radius: 50%; }

  /* ── Grids ── */
  .grid-2 { display: grid; grid-template-columns: 1fr 1fr;
             gap: 16px; margin-bottom: 24px; }
  .grid-3 { display: grid; grid-template-columns: 1fr 1fr 1fr;
             gap: 16px; margin-bottom: 24px; }

  /* ── Chart Cards ── */
  .chart-card { background: var(--card); border: 1px solid var(--border);
                border-radius: 8px; padding: 20px;
                box-shadow: 0 1px 4px rgba(28,53,94,0.06); }
  .chart-card h3 { font-size: 12px; font-weight: 700; margin-bottom: 6px;
                   color: var(--muted); text-transform: uppercase;
                   letter-spacing: 0.6px; border-bottom: 2px solid #FF5F00;
                   padding-bottom: 8px; }
  .chart-desc { font-size: 11px; color: #8ba0bf; font-style: italic;
                margin-bottom: 14px; line-height: 1.5; }
  .sxs-pair { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }
  .sxs-pair .pair-card { background:#F5F8FC; border:1px solid var(--border);
                         border-radius:6px; padding: 12px; }
  .sxs-pair .pair-card.pilot { border-top: 3px solid #1C355E; }
  .sxs-pair .pair-card.newcohort { border-top: 3px solid #FF5F00; }
  .sxs-pair .pair-card h5 { font-size: 10px; text-transform: uppercase;
                            letter-spacing: 0.6px; color: var(--muted);
                            margin-bottom: 8px; }

  /* ── Bar Charts ── */
  .bar-group { margin-bottom: 14px; }
  .bar-label { font-size: 12px; color: var(--text); margin-bottom: 5px;
               display: flex; justify-content: space-between; }
  .bar-track { height: 22px; background: #E4EAF3; border-radius: 6px;
               overflow: hidden; position: relative; }
  .bar-fill { height: 100%; border-radius: 6px; transition: width 0.8s ease;
              display: flex; align-items: center; padding-left: 8px;
              font-size: 10px; font-weight: 600; color: white; }

  /* ── Donut Charts ── */
  .donut-container { display: flex; align-items: center; gap: 24px; }
  .donut-legend { flex: 1; }
  .legend-item { display: flex; align-items: center; gap: 8px;
                 margin-bottom: 10px; font-size: 13px; }
  .legend-dot { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }
  .legend-value { margin-left: auto; font-weight: 600; font-size: 13px;
                  color: #1C355E; }

  /* ── Hotel Table ── */
  .table-container { background: var(--card); border: 1px solid var(--border);
                     border-radius: 8px; overflow: hidden; margin-bottom: 24px;
                     box-shadow: 0 1px 4px rgba(28,53,94,0.06); }
  .table-header-bar { padding: 14px 20px; display: flex;
                      justify-content: space-between; align-items: center;
                      border-bottom: 2px solid #FF5F00;
                      background: #1C355E; }
  .table-header-bar h3 { font-size: 12px; font-weight: 700; color: #FFFFFF;
                          text-transform: uppercase; letter-spacing: 0.5px; }
  .sort-controls { display: flex; gap: 6px; }
  .sort-btn { background: transparent; border: 1px solid rgba(255,255,255,0.4);
              color: rgba(255,255,255,0.8); font-size: 11px; padding: 4px 10px;
              border-radius: 6px; cursor: pointer; transition: all 0.2s;
              font-family: Arial, sans-serif; }
  .sort-btn:hover, .sort-btn.active { background: #FF5F00; color: white;
                                       border-color: #FF5F00; }

  table { width: 100%; border-collapse: collapse; }
  thead th { text-align: left; font-size: 9px; text-transform: uppercase;
             letter-spacing: 0.3px; color: #FFFFFF; font-weight: 700;
             padding: 6px 6px; border-bottom: 1px solid rgba(255,255,255,0.2);
             cursor: pointer; user-select: none; white-space: nowrap; }
  thead th:hover { color: #FFDD00; }
  thead th.sorted-asc::after  { content: ' ▲'; font-size: 8px; }
  thead th.sorted-desc::after { content: ' ▼'; font-size: 9px; }
  thead th.group-pms { border-bottom: 3px solid #8BAAD0; }
  thead th.group-web { border-bottom: 3px solid #7DC9A0; }
  thead th.group-summary { border-bottom: 3px solid #FF5F00; }
  thead tr { background: #1C355E; }
  thead tr.super-header th { padding: 5px 6px; font-size: 10px;
                              font-weight: 700; border-bottom: none; }
  thead tr.super-header th.pms-header { color: #FFDD00; text-align: center; }
  thead tr.super-header th.web-header { color: #90EEB8; text-align: center; }
  thead tr.super-header th.summary-header { color: #FF5F00; text-align: center; }
  tbody td { padding: 5px 6px; font-size: 11px;
             border-bottom: 1px solid var(--border); white-space: nowrap;
             color: #1C355E; }
  tbody tr:nth-child(even) { background: #F5F8FC; }
  tbody tr:hover { background: #E8F0FA; }
  tbody tr:last-child td { border-bottom: none; }
  .hotel-name { font-weight: 600; max-width: 200px; overflow: hidden;
                text-overflow: ellipsis; }
  .row-badge { display: inline-block; font-size: 9px; font-weight: 700;
               padding: 2px 7px; border-radius: 9px; margin-left: 6px;
               letter-spacing: 0.3px; vertical-align: 1px;
               text-transform: uppercase; }
  .row-badge.pilot { background: #1C355E; color: #FFFFFF; }
  .row-badge.newcohort { background: #FF5F00; color: #FFFFFF; }
  .num { text-align: right; font-variant-numeric: tabular-nums; }
  td.summary-col { background: rgba(255,95,0,0.04); }
  tbody tr:nth-child(even) td.summary-col { background: rgba(255,95,0,0.07); }
  td.pms-col { background: rgba(28,53,94,0.04); }
  tbody tr:nth-child(even) td.pms-col { background: rgba(28,53,94,0.07); }
  td.web-col { background: rgba(46,139,87,0.04); }
  tbody tr:nth-child(even) td.web-col { background: rgba(46,139,87,0.07); }
  .col-divider { border-left: 2px solid var(--border); }

  /* ── Tooltip ── */
  .tooltip { position: fixed; background: #FFFFFF;
             border: 1px solid var(--border); border-radius: 8px;
             padding: 10px 14px; font-size: 12px; pointer-events: none;
             z-index: 100; display: none; max-width: 260px;
             box-shadow: 0 6px 20px rgba(28,53,94,0.15);
             color: #1C355E; }

  /* ── Footer ── */
  .footer { text-align: center; padding: 20px;
            font-size: 11px; border-top: 3px solid #FF5F00;
            background: #1C355E; color: rgba(255,255,255,0.6); }

  /* ── Responsive ── */
  @media (max-width: 1100px) {
    .kpi-row { grid-template-columns: repeat(3, 1fr); }
    .grid-2, .grid-3, .sxs-wrap, .sxs-pair { grid-template-columns: 1fr; }
    .sxs-kpi-grid { grid-template-columns: 1fr 1fr 1fr; }
  }

</style>
</head>
<body>

<div class="brand-bar"></div>
<div class="header">
  <div style="margin-bottom:10px;">
  <a href="index.html" style="display:inline-flex;align-items:center;gap:6px;background:rgba(255,255,255,0.15);border:1px solid rgba(255,255,255,0.3);color:rgba(255,255,255,0.85);font-size:11px;font-weight:700;padding:6px 14px;border-radius:20px;text-decoration:none;transition:background 0.2s;font-family:Arial,sans-serif;" onmouseover="this.style.background='rgba(255,95,0,0.7)'" onmouseout="this.style.background='rgba(255,255,255,0.15)'">&#8592; Dashboard Portal</a>
  </div>
  <img src="__LOGO_URI__" alt="My Place Hotels" style="height:44px; display:block; margin-bottom:12px;">
  <h1>Member Rates Program — 2026 YTD <span class="pilot-badge">YTD</span></h1>
  <div class="subtitle">Year-to-date performance · __PERIOD_LABEL__ · 81 hotel properties (32 Pilot + 49 Non-Pilot) · All channels (PMS + Web)</div>
</div>

<!-- Sticky filter bar: cohort toggle + management-company dropdown -->
<div class="cohort-toggle-wrap">
  <div class="filter-bar">
    <div class="filter-group cohort-toggle">
      <span class="lbl">Cohort</span>
      <button class="cohort-btn active" data-mode="total">Total Portfolio <span class="ct-count">(__TOTAL_COUNT__)</span></button>
      <button class="cohort-btn" data-mode="pilot">Pilot Only <span class="ct-count">(__PILOT_COUNT__)</span></button>
      <button class="cohort-btn" data-mode="nonpilot">Non-Pilot Only <span class="ct-count">(__NONPILOT_COUNT__)</span></button>
      <button class="cohort-btn" data-mode="sxs">Side-by-Side</button>
    </div>
    <div class="filter-divider" aria-hidden="true"></div>
    <div class="filter-group">
      <span class="lbl">Mgmt Co</span>
      <span class="mgmt-select-wrap">
        <select id="mgmtSelect" class="mgmt-select" aria-label="Filter by Management Company"></select>
      </span>
    </div>
  </div>
</div>

<div class="dashboard">

  <!-- KPI CARDS / SXS KPI -->
  <div id="kpiZone"></div>

  <!-- CHANNEL & MEMBER TYPE CHARTS -->
  <div id="donutsZone"></div>

  <!-- METRIC COMPARISON BARS (ADR / RPB / LOS) -->
  <div id="metricsZone"></div>

  <!-- NEW MEMBER SIGN-UPS WIDGET -->
  <div id="newMemberZone" style="margin-bottom:24px;"></div>

  <!-- CANCEL RATE & NEW VS CURRENT -->
  <div id="cancelZone"></div>

  <!-- HOTEL TABLE -->
  <div class="table-container">
    <div class="table-header-bar">
      <div>
        <h3 id="tableTitle">Hotel-by-Hotel Breakdown — PMS &amp; Web by Member Type</h3>
        <p style="font-size:11px;color:rgba(255,255,255,0.55);font-style:italic;margin-top:3px;">Revenue, reservations, ADR, and length of stay for every property — sortable by any column.</p>
      </div>
      <div class="sort-controls">
        <button class="sort-btn active" onclick="sortTable('total_rev', event)">Total Revenue</button>
        <button class="sort-btn" onclick="sortTable('pms_cm_rev', event)">PMS CM Rev</button>
        <button class="sort-btn" onclick="sortTable('web_cm_rev', event)">Web CM Rev</button>
        <button class="sort-btn" onclick="sortTable('cohort', event)">By Cohort</button>
        <button class="sort-btn" onclick="sortTable('name', event)">A–Z</button>
      </div>
    </div>
    <div style="overflow-x:auto;">
      <table>
        <thead>
          <tr class="super-header">
            <th rowspan="2" onclick="sortTable('name', event)" style="vertical-align:bottom;">Property</th>
            <th rowspan="2" onclick="sortTable('cohort', event)" style="vertical-align:bottom;">Cohort</th>
            <th rowspan="2" onclick="sortTable('guests', event)" class="num" style="vertical-align:bottom;">Guests</th>
            <th colspan="4" class="pms-header col-divider">PMS (Property)</th>
            <th colspan="4" class="web-header col-divider">WEB (Website)</th>
            <th colspan="4" class="summary-header col-divider">Totals</th>
          </tr>
          <tr>
            <th onclick="sortTable('pms_cm_rev', event)" class="num group-pms col-divider">CM Rev</th>
            <th onclick="sortTable('pms_nm_rev', event)" class="num group-pms">NM Rev</th>
            <th onclick="sortTable('pms_cm_res', event)" class="num group-pms">CM Res</th>
            <th onclick="sortTable('pms_nm_res', event)" class="num group-pms">NM Res</th>
            <th onclick="sortTable('web_cm_rev', event)" class="num group-web col-divider">CM Rev</th>
            <th onclick="sortTable('web_nm_rev', event)" class="num group-web">NM Rev</th>
            <th onclick="sortTable('web_cm_res', event)" class="num group-web">CM Res</th>
            <th onclick="sortTable('web_nm_res', event)" class="num group-web">NM Res</th>
            <th onclick="sortTable('total_rev', event)" class="num group-summary col-divider">Revenue</th>
            <th onclick="sortTable('total_res', event)" class="num group-summary">Res</th>
            <th onclick="sortTable('net_adr', event)" class="num group-summary">ADR</th>
            <th onclick="sortTable('rev_per_res', event)" class="num group-summary">RpR</th>
          </tr>
        </thead>
        <tbody id="hotelTableBody"></tbody>
      </table>
    </div>
  </div>

</div>

<div class="footer">
  Member Rates Program Dashboard · Data as of __DATA_AS_OF__ · Rate codes: MEMBER RATE, MEMBER RATE EXT, MEMBER RATE HOTEL, MEMBER RATE EXT HOTEL · Excludes cancelled reservations
</div>

<div class="tooltip" id="tooltip"></div>

<script>
// ===== DATA =====
const DASH_DATA = __DATA_JSON__;
const COHORTS = DASH_DATA.cohorts;       // {total, pilot, nonpilot} — baseline Python-side aggs
const COUNTS = DASH_DATA.counts;         // {total, pilot, nonpilot}
const HOTELS_ALL = DASH_DATA.hotels;     // every hotel (immutable canonical dataset)
const MGMT_OPTIONS = DASH_DATA.mgmtOptions; // [{name, count}, ...] with 'All Companies' pinned first
// Pre-compute table-friendly totals once
HOTELS_ALL.forEach(h => {
  h.total_rev = h.pms_cm_rev + h.pms_nm_rev + h.web_cm_rev + h.web_nm_rev;
  h.total_res = h.pms_cm_res + h.pms_nm_res + h.web_cm_res + h.web_nm_res;
  h.rev_per_res = h.total_res > 0 ? h.total_rev / h.total_res : 0;
});

const ALL_MGMT = 'All Companies';
let activeMode = 'total';        // 'total' | 'pilot' | 'nonpilot' | 'sxs'
let activeMgmt = ALL_MGMT;       // operator name or 'All Companies'
let currentSort = 'total_rev';
let sortDir = 'desc';

// ===== AGGREGATION (mirrors the Python aggregate() shape) =====
// Sums the per-hotel `cells` raw fields across the supplied hotel array,
// then derives ADR/RPB/LOS/cancel_rate from the totals — exactly like the
// Python aggregate() function does. Result is a drop-in replacement for the
// pre-computed COHORTS[*] objects, so every existing render fn works
// unchanged with the dynamic agg.
const RAW_FIELDS = ['guests','gross_res','net_res','gross_rn','net_rn','gross_rev','net_rev'];
const MT_BUCKETS = ['current','new_mem','non_mem','unknown'];

function _emptyBucket() {
  const o = {};
  for (const f of RAW_FIELDS) o[f] = 0;
  return o;
}
function _addInto(dst, src) {
  for (const f of RAW_FIELDS) dst[f] += (src[f] || 0);
}
function _derive(b) {
  const out = Object.assign({}, b);
  out.net_adr     = b.net_rn  ? (b.net_rev / b.net_rn)  : 0;
  out.gross_adr   = b.gross_rn ? (b.gross_rev / b.gross_rn) : 0;
  out.net_rpb     = b.net_res ? (b.net_rev / b.net_res) : 0;
  out.net_los     = b.net_res ? (b.net_rn  / b.net_res) : 0;
  out.cancel_rate = b.gross_res ? (1 - (b.net_res / b.gross_res)) : 0;
  return out;
}
function aggregateFromHotels(hotels) {
  const pmsBuckets = {}, webBuckets = {};
  for (const k of MT_BUCKETS) { pmsBuckets[k] = _emptyBucket(); webBuckets[k] = _emptyBucket(); }
  for (const h of hotels) {
    if (!h.cells) continue;
    for (const k of MT_BUCKETS) {
      if (h.cells.pms && h.cells.pms[k]) _addInto(pmsBuckets[k], h.cells.pms[k]);
      if (h.cells.web && h.cells.web[k]) _addInto(webBuckets[k], h.cells.web[k]);
    }
  }
  function foldChan(buckets) {
    const tot = _emptyBucket();
    for (const k of MT_BUCKETS) _addInto(tot, buckets[k]);
    return {
      total:   _derive(tot),
      current: _derive(buckets.current),
      new_mem: _derive(buckets.new_mem),
      non_mem: _derive(buckets.non_mem),
    };
  }
  const pms = foldChan(pmsBuckets);
  const web = foldChan(webBuckets);
  const grand = _emptyBucket();
  for (const f of RAW_FIELDS) grand[f] = pms.total[f] + web.total[f];
  const grandD = _derive(grand);
  return { pms, web, grand: grandD };
}

// ===== FILTER PIPELINE =====
function matchesMgmt(h) {
  return activeMgmt === ALL_MGMT || h.mgmt_co === activeMgmt;
}
function filterByMode(hotels, mode) {
  if (mode === 'pilot')    return hotels.filter(h => h.pilot);
  if (mode === 'nonpilot') return hotels.filter(h => !h.pilot);
  return hotels;
}
// Final filtered hotel set for the current view (mgmt always applied;
// cohort applied unless in side-by-side mode, where each side filters
// independently).
function activeHotels() {
  const m = HOTELS_ALL.filter(matchesMgmt);
  if (activeMode === 'sxs') return m;            // SXS does its own pilot split
  return filterByMode(m, activeMode);
}
function activeCohortLabel() {
  if (activeMode === 'pilot')    return 'Pilot Cohort';
  if (activeMode === 'nonpilot') return 'Non-Pilot Cohort';
  if (activeMode === 'sxs')      return 'All Properties (Pilot + Non-Pilot)';
  return 'Total Portfolio';
}
function mgmtSuffix() {
  if (activeMgmt === ALL_MGMT) return '';
  const c = HOTELS_ALL.filter(matchesMgmt).length;
  return ' · ' + activeMgmt + ' (' + c + ' hotel' + (c === 1 ? '' : 's') + ')';
}

// ===== UTILS =====
function fmt(n) {
  if (n === null || n === undefined || isNaN(n)) return '0';
  if (n >= 1000) return n.toLocaleString('en-US', {maximumFractionDigits:0});
  return Math.round(n).toString();
}
function fmtMoney(n) { return '$' + fmt(n); }
function pct(n) { return (n*100).toFixed(1) + '%'; }

// ===== KPI RENDERING =====
function renderKPISingle(agg) {
  const g = agg.grand;
  const kpis = [
    {label:'Net Revenue', value:fmtMoney(g.net_rev), sub:`<span class="pms">PMS ${fmtMoney(agg.pms.total.net_rev)}</span> · <span class="web">Web ${fmtMoney(agg.web.total.net_rev)}</span>`},
    {label:'Net Reservations', value:fmt(g.net_res), sub:`<span class="pms">PMS ${fmt(agg.pms.total.net_res)}</span> · <span class="web">Web ${fmt(agg.web.total.net_res)}</span>`},
    {label:'Total Guests', value:fmt(g.guests), sub:`<span class="pms">PMS ${fmt(agg.pms.total.guests)}</span> · <span class="web">Web ${fmt(agg.web.total.guests)}</span>`},
    {label:'Net ADR', value:'$'+g.net_adr.toFixed(2), sub:`<span class="pms">PMS $${agg.pms.total.net_adr.toFixed(2)}</span> · <span class="web">Web $${agg.web.total.net_adr.toFixed(2)}</span>`},
    {label:'Rev per Reservation', value:'$'+g.net_rpb.toFixed(2), sub:`<span class="pms">PMS $${agg.pms.total.net_rpb.toFixed(2)}</span> · <span class="web">Web $${agg.web.total.net_rpb.toFixed(2)}</span>`},
    {label:'Cancel Rate', value:pct(g.cancel_rate), sub:`<span class="pms">PMS ${pct(agg.pms.total.cancel_rate)}</span> · <span class="web">Web ${pct(agg.web.total.cancel_rate)}</span>`}
  ];
  return `<div class="kpi-row">${kpis.map(k=>`
    <div class="kpi-card">
      <div class="kpi-label">${k.label}</div>
      <div class="kpi-value">${k.value}</div>
      <div class="kpi-sub">${k.sub}</div>
    </div>`).join('')}</div>`;
}

// Resolves the {pilot, nonpilot} aggs used by side-by-side mode, honoring
// the current Mgmt-Co filter. Each side is the intersection of (pilot flag) ∩ (mgmt).
function sxsAggs() {
  const m = HOTELS_ALL.filter(matchesMgmt);
  return {
    pilot:    aggregateFromHotels(m.filter(h => h.pilot)),
    nonpilot: aggregateFromHotels(m.filter(h => !h.pilot)),
    pilotCount:    m.filter(h => h.pilot).length,
    nonpilotCount: m.filter(h => !h.pilot).length,
  };
}

function renderKPISideBySide() {
  const sx = sxsAggs();
  const cols = [
    {key:'pilot', cls:'pilot', label:'Pilot Cohort', tag:'PILOT', count: sx.pilotCount, agg: sx.pilot},
    {key:'nonpilot', cls:'newcohort', label:'Non-Pilot Cohort', tag:'NEW', count: sx.nonpilotCount, agg: sx.nonpilot}
  ];
  return `<div class="sxs-wrap">${cols.map(c => {
    const a = c.agg;
    const g = a.grand;
    return `<div class="sxs-col ${c.cls}">
      <h4>${c.label} <span class="tag">${c.tag}</span> <span style="font-size:11px;color:var(--muted);font-weight:normal;">(${c.count} hotels)</span></h4>
      <div class="sxs-kpi-grid">
        <div class="sxs-kpi"><div class="lbl">Net Revenue</div><div class="val">${fmtMoney(g.net_rev)}</div><div class="sub">PMS ${fmtMoney(a.pms.total.net_rev)} · Web ${fmtMoney(a.web.total.net_rev)}</div></div>
        <div class="sxs-kpi"><div class="lbl">Net Reservations</div><div class="val">${fmt(g.net_res)}</div><div class="sub">PMS ${fmt(a.pms.total.net_res)} · Web ${fmt(a.web.total.net_res)}</div></div>
        <div class="sxs-kpi"><div class="lbl">Net ADR</div><div class="val">$${g.net_adr.toFixed(2)}</div><div class="sub">PMS $${a.pms.total.net_adr.toFixed(2)} · Web $${a.web.total.net_adr.toFixed(2)}</div></div>
        <div class="sxs-kpi"><div class="lbl">Total Guests</div><div class="val">${fmt(g.guests)}</div><div class="sub">PMS ${fmt(a.pms.total.guests)} · Web ${fmt(a.web.total.guests)}</div></div>
        <div class="sxs-kpi"><div class="lbl">Rev / Booking</div><div class="val">$${g.net_rpb.toFixed(2)}</div><div class="sub">PMS $${a.pms.total.net_rpb.toFixed(2)} · Web $${a.web.total.net_rpb.toFixed(2)}</div></div>
        <div class="sxs-kpi"><div class="lbl">Cancel Rate</div><div class="val">${pct(g.cancel_rate)}</div><div class="sub">PMS ${pct(a.pms.total.cancel_rate)} · Web ${pct(a.web.total.cancel_rate)}</div></div>
      </div>
    </div>`;
  }).join('')}</div>`;
}

// ===== DONUTS =====
function drawDonut(canvas, data, colors, center_top, center_bot) {
  const ctx = canvas.getContext('2d');
  ctx.clearRect(0, 0, canvas.width, canvas.height);
  const cx = canvas.width/2, cy = canvas.height/2;
  const r = Math.min(canvas.width, canvas.height)/2 - 20;
  const inner = r * 0.76;
  const total = data.reduce((s,d)=>s+d.value,0);
  let start = -Math.PI/2;
  data.forEach((d,i) => {
    const angle = (d.value/total)*Math.PI*2;
    ctx.beginPath();
    ctx.arc(cx, cy, r, start, start+angle);
    ctx.arc(cx, cy, inner, start+angle, start, true);
    ctx.closePath();
    ctx.fillStyle = colors[i];
    ctx.fill();
    start += angle;
  });
  ctx.fillStyle = '#1C355E';
  ctx.font = 'bold 20px Arial';
  ctx.textAlign = 'center';
  ctx.textBaseline = 'middle';
  ctx.fillText(center_top, cx, cy-8);
  ctx.font = '10px Arial';
  ctx.fillStyle = '#5A6B84';
  ctx.fillText(center_bot, cx, cy+12);
}

function donutCardHTML(id_canvas, id_legend, title, desc) {
  return `<div class="chart-card">
    <h3>${title}</h3>
    <p class="chart-desc">${desc}</p>
    <div class="donut-container">
      <canvas id="${id_canvas}" width="220" height="220"></canvas>
      <div class="donut-legend" id="${id_legend}"></div>
    </div>
  </div>`;
}

function fillChannelDonut(agg, canvasId, legendId) {
  const data = [
    {label:'PMS (Property)', value: agg.pms.total.net_rev},
    {label:'Web', value: agg.web.total.net_rev}
  ];
  const colors = ['#1C355E','#2E8B57'];
  drawDonut(document.getElementById(canvasId), data, colors, '$'+fmt(data[0].value+data[1].value), 'NET REVENUE');
  const total = data.reduce((s,d)=>s+d.value,0);
  document.getElementById(legendId).innerHTML = data.map((d,i)=>`
    <div class="legend-item">
      <div class="legend-dot" style="background:${colors[i]}"></div>
      <span>${d.label}</span>
      <span class="legend-value">$${fmt(d.value)} (${(d.value/total*100).toFixed(1)}%)</span>
    </div>`).join('');
}

function fillMemberDonut(agg, canvasId, legendId) {
  const curr = agg.pms.current.guests + agg.web.current.guests;
  const newm = agg.pms.new_mem.guests + agg.web.new_mem.guests;
  const data = [
    {label:'Current Members', value: curr},
    {label:'New Members', value: newm}
  ];
  const colors = ['#1C355E','#2E8B57'];
  drawDonut(document.getElementById(canvasId), data, colors, fmt(curr+newm), 'NET GUESTS');
  const total = data.reduce((s,d)=>s+d.value,0);
  document.getElementById(legendId).innerHTML = data.map((d,i)=>`
    <div class="legend-item">
      <div class="legend-dot" style="background:${colors[i]}"></div>
      <span>${d.label}</span>
      <span class="legend-value">${fmt(d.value)} (${(d.value/total*100).toFixed(1)}%)</span>
    </div>`).join('');
}

function renderDonutsSingle(agg) {
  document.getElementById('donutsZone').innerHTML = `<div class="grid-2">
    ${donutCardHTML('channelDonut','channelLegend','Net Revenue by Channel','Revenue split between PMS (property-booked) and Web (website-booked) booking channels.')}
    ${donutCardHTML('memberDonut','memberLegend','Net Guests by Member Type','Guest share split between returning Current Members and newly enrolled New Members.')}
  </div>`;
  fillChannelDonut(agg, 'channelDonut', 'channelLegend');
  fillMemberDonut(agg, 'memberDonut', 'memberLegend');
}

function renderDonutsSXS() {
  const sx = sxsAggs();
  document.getElementById('donutsZone').innerHTML = `
    <div class="chart-card" style="margin-bottom:18px;">
      <h3>Net Revenue by Channel — Pilot vs Non-Pilot</h3>
      <p class="chart-desc">Revenue split between PMS and Web for each cohort side by side.</p>
      <div class="sxs-pair">
        <div class="pair-card pilot"><h5>Pilot Cohort</h5>
          <div class="donut-container"><canvas id="cd_pilot" width="180" height="180"></canvas><div class="donut-legend" id="cl_pilot"></div></div></div>
        <div class="pair-card newcohort"><h5>Non-Pilot Cohort</h5>
          <div class="donut-container"><canvas id="cd_np" width="180" height="180"></canvas><div class="donut-legend" id="cl_np"></div></div></div>
      </div>
    </div>
    <div class="chart-card" style="margin-bottom:24px;">
      <h3>Net Guests by Member Type — Pilot vs Non-Pilot</h3>
      <p class="chart-desc">Returning vs new member mix for each cohort.</p>
      <div class="sxs-pair">
        <div class="pair-card pilot"><h5>Pilot Cohort</h5>
          <div class="donut-container"><canvas id="md_pilot" width="180" height="180"></canvas><div class="donut-legend" id="ml_pilot"></div></div></div>
        <div class="pair-card newcohort"><h5>Non-Pilot Cohort</h5>
          <div class="donut-container"><canvas id="md_np" width="180" height="180"></canvas><div class="donut-legend" id="ml_np"></div></div></div>
      </div>
    </div>`;
  fillChannelDonut(sx.pilot,    'cd_pilot', 'cl_pilot');
  fillChannelDonut(sx.nonpilot, 'cd_np',    'cl_np');
  fillMemberDonut(sx.pilot,     'md_pilot', 'ml_pilot');
  fillMemberDonut(sx.nonpilot,  'md_np',    'ml_np');
}

// ===== METRIC BARS (ADR / RPB / LOS) =====
function barsHTML(items, maxVal, prefix, suffix) {
  const colors = {'PMS — Current Member':'#1C355E','PMS — New Member':'#4A72A8',
                  'Web — Current Member':'#2E8B57','Web — New Member':'#5AB87E'};
  return items.map(d => {
    const pct = Math.min(100, (d.value/maxVal*100)).toFixed(1);
    const color = colors[d.label] || '#1C355E';
    return `<div class="bar-group">
      <div class="bar-label"><span>${d.label}</span><span>${prefix||''}${(+d.value).toFixed(2)}${suffix||''}</span></div>
      <div class="bar-track"><div class="bar-fill" style="width:${pct}%;background:${color}"></div></div>
    </div>`;
  }).join('');
}

function metricItems(agg, field) {
  return [
    {label:'PMS — Current Member', value: agg.pms.current[field]},
    {label:'PMS — New Member',     value: agg.pms.new_mem[field]},
    {label:'Web — Current Member', value: agg.web.current[field]},
    {label:'Web — New Member',     value: agg.web.new_mem[field]}
  ];
}

function renderMetricsSingle(agg) {
  document.getElementById('metricsZone').innerHTML = `<div class="grid-3">
    <div class="chart-card"><h3>ADR Comparison</h3>
      <p class="chart-desc">Average Daily Rate by channel and member type, showing rate differences between booking paths.</p>
      <div>${barsHTML(metricItems(agg,'net_adr'),200,'$','')}</div></div>
    <div class="chart-card"><h3>Revenue per Booking</h3>
      <p class="chart-desc">Average net revenue per reservation across channels and member types — reflects stay length and rate combined.</p>
      <div>${barsHTML(metricItems(agg,'net_rpb'),500,'$','')}</div></div>
    <div class="chart-card"><h3>Avg Length of Stay</h3>
      <p class="chart-desc">Mean number of nights per stay, segmented by booking channel and member type.</p>
      <div>${barsHTML(metricItems(agg,'net_los'),5,'',' nights')}</div></div>
  </div>`;
}

function renderMetricsSXS() {
  const sx = sxsAggs();
  function pair(field, prefix, suffix, maxV, title, desc) {
    return `<div class="chart-card"><h3>${title}</h3><p class="chart-desc">${desc}</p>
      <div class="sxs-pair">
        <div class="pair-card pilot"><h5>Pilot Cohort</h5>${barsHTML(metricItems(sx.pilot,field),maxV,prefix,suffix)}</div>
        <div class="pair-card newcohort"><h5>Non-Pilot Cohort</h5>${barsHTML(metricItems(sx.nonpilot,field),maxV,prefix,suffix)}</div>
      </div></div>`;
  }
  document.getElementById('metricsZone').innerHTML = `<div class="grid-3">
    ${pair('net_adr','$','',200,'ADR Comparison','Average Daily Rate by channel and member type.')}
    ${pair('net_rpb','$','',500,'Revenue per Booking','Average net revenue per reservation.')}
    ${pair('net_los','',' nights',5,'Avg Length of Stay','Mean number of nights per stay.')}
  </div>`;
}

// ===== NEW MEMBER SIGN-UPS WIDGET =====
function newMemberWidgetHTML(agg, cohortLabel) {
  const p = agg.pms.new_mem, w = agg.web.new_mem;
  const totalSignups = p.guests + w.guests;
  const totalRes = p.net_res + w.net_res;
  const totalRev = p.net_rev + w.net_rev;
  const pmsPct = totalSignups > 0 ? (p.guests/totalSignups*100).toFixed(1) : '0';
  const webPct = totalSignups > 0 ? (w.guests/totalSignups*100).toFixed(1) : '0';
  return `<div class="chart-card">
    <h3>New Member Sign-Ups by Channel${cohortLabel ? ' — '+cohortLabel : ''}</h3>
    <p class="chart-desc">Total new member enrollments generated through each booking channel during the YTD period.</p>
    <div style="display:grid;grid-template-columns:1fr 2fr 1fr;gap:24px;align-items:center;">
      <div style="text-align:center;padding:16px;background:#E4EAF3;border-radius:10px;border:1px solid var(--border);">
        <div style="font-size:11px;text-transform:uppercase;letter-spacing:0.8px;color:var(--accent);font-weight:600;margin-bottom:8px;">PMS (Property)</div>
        <div style="font-size:36px;font-weight:700;color:var(--accent);letter-spacing:-1px;">${fmt(p.guests)}</div>
        <div style="font-size:12px;color:var(--muted);margin-top:4px;">New Member Sign-Ups</div>
        <div style="margin-top:12px;display:flex;justify-content:space-around;">
          <div style="text-align:center;"><div style="font-size:14px;font-weight:600;">${fmt(p.net_res)}</div><div style="font-size:10px;color:var(--muted);">Reservations</div></div>
          <div style="text-align:center;"><div style="font-size:14px;font-weight:600;">${fmtMoney(p.net_rev)}</div><div style="font-size:10px;color:var(--muted);">Net Revenue</div></div>
        </div>
      </div>
      <div style="text-align:center;">
        <div style="font-size:11px;text-transform:uppercase;letter-spacing:0.8px;color:var(--muted);font-weight:600;margin-bottom:6px;">Total New Member Sign-Ups</div>
        <div style="font-size:48px;font-weight:700;letter-spacing:-2px;background:linear-gradient(135deg,var(--accent),var(--accent2));-webkit-background-clip:text;-webkit-text-fill-color:transparent;">${fmt(totalSignups)}</div>
        <div style="font-size:13px;color:var(--muted);margin-top:2px;">${fmt(totalRes)} reservations · ${fmtMoney(totalRev)} net revenue</div>
        <div style="margin-top:14px;">
          <div style="height:12px;border-radius:6px;overflow:hidden;display:flex;background:#D0D8E8;">
            <div style="width:${pmsPct}%;background:var(--accent);transition:width 0.8s ease;"></div>
            <div style="width:${webPct}%;background:var(--accent2);transition:width 0.8s ease;"></div>
          </div>
          <div style="display:flex;justify-content:space-between;font-size:10px;color:var(--muted);margin-top:4px;">
            <span style="color:var(--accent);">▌ PMS ${pmsPct}%</span>
            <span style="color:var(--accent2);">WEB ${webPct}% ▌</span>
          </div>
        </div>
      </div>
      <div style="text-align:center;padding:16px;background:#E4EAF3;border-radius:10px;border:1px solid var(--border);">
        <div style="font-size:11px;text-transform:uppercase;letter-spacing:0.8px;color:var(--accent2);font-weight:600;margin-bottom:8px;">WEB (Website)</div>
        <div style="font-size:36px;font-weight:700;color:var(--accent2);letter-spacing:-1px;">${fmt(w.guests)}</div>
        <div style="font-size:12px;color:var(--muted);margin-top:4px;">New Member Sign-Ups</div>
        <div style="margin-top:12px;display:flex;justify-content:space-around;">
          <div style="text-align:center;"><div style="font-size:14px;font-weight:600;">${fmt(w.net_res)}</div><div style="font-size:10px;color:var(--muted);">Reservations</div></div>
          <div style="text-align:center;"><div style="font-size:14px;font-weight:600;">${fmtMoney(w.net_rev)}</div><div style="font-size:10px;color:var(--muted);">Net Revenue</div></div>
        </div>
      </div>
    </div></div>`;
}

function renderNewMemberSingle(agg) {
  document.getElementById('newMemberZone').innerHTML = newMemberWidgetHTML(agg, null);
}
function renderNewMemberSXS() {
  const sx = sxsAggs();
  document.getElementById('newMemberZone').innerHTML = `
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;">
      ${newMemberWidgetHTML(sx.pilot,    'Pilot Cohort')}
      ${newMemberWidgetHTML(sx.nonpilot, 'Non-Pilot Cohort')}
    </div>`;
}

// ===== CANCEL RATE & NEW vs CURRENT =====
function cancelItems(agg) {
  return [
    {label:'PMS — Current Member', value: agg.pms.current.cancel_rate*100},
    {label:'PMS — New Member',     value: agg.pms.new_mem.cancel_rate*100},
    {label:'Web — Current Member', value: agg.web.current.cancel_rate*100},
    {label:'Web — New Member',     value: agg.web.new_mem.cancel_rate*100}
  ];
}
function cancelBarsHTML(items, maxVal) {
  const colors = {'PMS — Current Member':'#1C355E','PMS — New Member':'#4A72A8','Web — Current Member':'#2E8B57','Web — New Member':'#5AB87E'};
  return items.map(d => {
    const pct = Math.min(100, (d.value/maxVal*100)).toFixed(1);
    const color = colors[d.label] || '#1C355E';
    return `<div class="bar-group">
      <div class="bar-label"><span>${d.label}</span><span>${d.value.toFixed(1)}%</span></div>
      <div class="bar-track"><div class="bar-fill" style="width:${pct}%;background:${color}"></div></div>
    </div>`;
  }).join('');
}

function newVsCurrentHTML(agg) {
  const currRev = agg.pms.current.net_rev + agg.web.current.net_rev;
  const newRev  = agg.pms.new_mem.net_rev + agg.web.new_mem.net_rev;
  const currRes = agg.pms.current.net_res + agg.web.current.net_res;
  const newRes  = agg.pms.new_mem.net_res + agg.web.new_mem.net_res;
  const currGuests = agg.pms.current.guests + agg.web.current.guests;
  const newGuests  = agg.pms.new_mem.guests + agg.web.new_mem.guests;
  function pair(label, c, n) {
    const tot = c + n;
    const cp = tot ? c/tot*100 : 0;
    const np = tot ? n/tot*100 : 0;
    const fmtVal = (v) => (label==='Net Revenue') ? fmtMoney(v) : fmt(v);
    return `<div class="bar-group">
      <div class="bar-label"><span>${label}</span><span style="color:var(--accent)">Current: ${fmtVal(c)}</span></div>
      <div class="bar-track" style="display:flex">
        <div style="width:${cp}%;background:#1C355E;height:100%;border-radius:6px 0 0 6px;"></div>
        <div style="width:${np}%;background:#2E8B57;height:100%;border-radius:0 6px 6px 0;"></div>
      </div>
      <div style="display:flex;justify-content:space-between;font-size:10px;color:var(--muted);margin-top:2px;">
        <span>Current ${cp.toFixed(0)}%</span>
        <span style="color:var(--accent2)">New: ${fmtVal(n)} (${np.toFixed(0)}%)</span>
      </div></div>`;
  }
  return pair('Net Revenue', currRev, newRev) + pair('Net Reservations', currRes, newRes) + pair('Guests', currGuests, newGuests);
}

function renderCancelSingle(agg) {
  document.getElementById('cancelZone').innerHTML = `<div class="grid-2">
    <div class="chart-card"><h3>Cancellation Rate by Channel &amp; Type</h3>
      <p class="chart-desc">Percentage of reservations that were cancelled, broken down by booking channel and member type.</p>
      ${cancelBarsHTML(cancelItems(agg), 25)}
    </div>
    <div class="chart-card"><h3>New Member vs Current Member — Key Metrics</h3>
      <p class="chart-desc">Side-by-side comparison of revenue, reservations, and guests between newly enrolled and returning members.</p>
      ${newVsCurrentHTML(agg)}
    </div>
  </div>`;
}

function renderCancelSXS() {
  const sx = sxsAggs();
  document.getElementById('cancelZone').innerHTML = `<div class="grid-2">
    <div class="chart-card"><h3>Cancellation Rate — Pilot vs Non-Pilot</h3>
      <p class="chart-desc">Cancellation by channel and member type, each cohort.</p>
      <div class="sxs-pair">
        <div class="pair-card pilot"><h5>Pilot Cohort</h5>${cancelBarsHTML(cancelItems(sx.pilot), 25)}</div>
        <div class="pair-card newcohort"><h5>Non-Pilot Cohort</h5>${cancelBarsHTML(cancelItems(sx.nonpilot), 25)}</div>
      </div>
    </div>
    <div class="chart-card"><h3>New vs Current — Pilot vs Non-Pilot</h3>
      <p class="chart-desc">Revenue / reservation / guest mix between new and current members, each cohort.</p>
      <div class="sxs-pair">
        <div class="pair-card pilot"><h5>Pilot Cohort</h5>${newVsCurrentHTML(sx.pilot)}</div>
        <div class="pair-card newcohort"><h5>Non-Pilot Cohort</h5>${newVsCurrentHTML(sx.nonpilot)}</div>
      </div>
    </div>
  </div>`;
}

// ===== HOTEL TABLE =====
function activeAggForTotalsRow() {
  // Always recompute from the canonical filtered hotel set so both filters
  // (cohort + mgmt-co) stack consistently across every render path.
  return aggregateFromHotels(activeHotels());
}

function sortTable(field, evt) {
  if (evt && evt.target && evt.target.classList && evt.target.classList.contains('sort-btn')) {
    document.querySelectorAll('.sort-controls .sort-btn').forEach(b => b.classList.remove('active'));
    evt.target.classList.add('active');
  }
  if (currentSort === field) sortDir = sortDir === 'desc' ? 'asc' : 'desc';
  else { currentSort = field; sortDir = (field === 'name' || field === 'cohort') ? 'asc' : 'desc'; }
  renderHotelTable();
}

function renderHotelTable() {
  const list = activeHotels();
  if (list.length === 0) {
    // Empty intersection: show a single message row spanning all columns.
    document.getElementById('hotelTableBody').innerHTML =
      `<tr><td colspan="15" style="text-align:center;padding:30px 20px;color:var(--muted);font-style:italic;">
         No properties match this filter combination.
       </td></tr>`;
    return;
  }
  const sorted = [...list].sort((a,b) => {
    let va = a[currentSort], vb = b[currentSort];
    if (currentSort === 'name' || currentSort === 'cohort') {
      return sortDir === 'asc' ? String(va).localeCompare(String(vb)) : String(vb).localeCompare(String(va));
    }
    return sortDir === 'desc' ? (vb - va) : (va - vb);
  });

  // Totals row uses cohort-aggregated numbers (authoritative)
  const agg = activeAggForTotalsRow();
  const g = agg.grand;
  const t = {
    guests: g.guests,
    pms_cm_rev: agg.pms.current.net_rev,
    pms_nm_rev: agg.pms.new_mem.net_rev,
    pms_cm_res: agg.pms.current.net_res,
    pms_nm_res: agg.pms.new_mem.net_res,
    web_cm_rev: agg.web.current.net_rev,
    web_nm_rev: agg.web.new_mem.net_rev,
    web_cm_res: agg.web.current.net_res,
    web_nm_res: agg.web.new_mem.net_res,
    total_rev: g.net_rev,
    total_res: g.net_res
  };
  const t_adr = g.net_adr;
  const t_rpr = g.net_rpb;

  document.getElementById('hotelTableBody').innerHTML = sorted.map(h => {
    const badgeCls = h.pilot ? 'pilot' : 'newcohort';
    const badgeText = h.pilot ? 'Pilot' : 'New';
    return `<tr>
      <td class="hotel-name">${h.name}<span class="row-badge ${badgeCls}">${badgeText}</span></td>
      <td><span class="row-badge ${badgeCls}">${badgeText}</span></td>
      <td class="num">${fmt(h.guests)}</td>
      <td class="num pms-col col-divider">$${fmt(h.pms_cm_rev)}</td>
      <td class="num pms-col">$${fmt(h.pms_nm_rev)}</td>
      <td class="num pms-col">${fmt(h.pms_cm_res)}</td>
      <td class="num pms-col">${fmt(h.pms_nm_res)}</td>
      <td class="num web-col col-divider">$${fmt(h.web_cm_rev)}</td>
      <td class="num web-col">$${fmt(h.web_nm_rev)}</td>
      <td class="num web-col">${fmt(h.web_cm_res)}</td>
      <td class="num web-col">${fmt(h.web_nm_res)}</td>
      <td class="num summary-col col-divider" style="font-weight:600">$${fmt(h.total_rev)}</td>
      <td class="num summary-col">${fmt(h.total_res)}</td>
      <td class="num summary-col">$${h.net_adr.toFixed(2)}</td>
      <td class="num summary-col" style="font-weight:600">$${h.rev_per_res.toFixed(2)}</td>
    </tr>`;
  }).join('') + `<tr style="font-weight:700;background:#1C355E;border-top:2px solid #FF5F00;">
      <td style="color:#fff;">TOTALS</td>
      <td style="color:#fff;">—</td>
      <td class="num" style="color:#fff;">${fmt(t.guests)}</td>
      <td class="num col-divider" style="color:#fff;">$${fmt(t.pms_cm_rev)}</td>
      <td class="num" style="color:#fff;">$${fmt(t.pms_nm_rev)}</td>
      <td class="num" style="color:#fff;">${fmt(t.pms_cm_res)}</td>
      <td class="num" style="color:#fff;">${fmt(t.pms_nm_res)}</td>
      <td class="num col-divider" style="color:#fff;">$${fmt(t.web_cm_rev)}</td>
      <td class="num" style="color:#fff;">$${fmt(t.web_nm_rev)}</td>
      <td class="num" style="color:#fff;">${fmt(t.web_cm_res)}</td>
      <td class="num" style="color:#fff;">${fmt(t.web_nm_res)}</td>
      <td class="num col-divider" style="color:#fff;">$${fmt(t.total_rev)}</td>
      <td class="num" style="color:#fff;">${fmt(t.total_res)}</td>
      <td class="num" style="color:#fff;">$${t_adr.toFixed(2)}</td>
      <td class="num" style="color:#fff;">$${t_rpr.toFixed(2)}</td>
    </tr>`;
}

// ===== TABLE TITLE UPDATE =====
function updateTableTitle() {
  const el = document.getElementById('tableTitle');
  const base = 'Hotel-by-Hotel Breakdown — PMS & Web by Member Type';
  let cohortLabel;
  if (activeMode === 'pilot') {
    const ct = HOTELS_ALL.filter(matchesMgmt).filter(h => h.pilot).length;
    cohortLabel = ' — Pilot Cohort (' + ct + ')';
  } else if (activeMode === 'nonpilot') {
    const ct = HOTELS_ALL.filter(matchesMgmt).filter(h => !h.pilot).length;
    cohortLabel = ' — Non-Pilot Cohort (' + ct + ')';
  } else if (activeMode === 'sxs') {
    cohortLabel = ' — All Properties (Pilot + Non-Pilot)';
  } else {
    const ct = HOTELS_ALL.filter(matchesMgmt).length;
    cohortLabel = ' — Total Portfolio (' + ct + ')';
  }
  el.textContent = base + cohortLabel + mgmtSuffix();
}

// ===== EMPTY-STATE =====
function emptyStateHTML(message) {
  return `<div class="empty-state">
    <div class="es-title">No properties match this filter combination</div>
    <div class="es-sub">${message}</div>
  </div>`;
}
function renderEmptyZones(message) {
  const note = emptyStateHTML(message);
  document.getElementById('kpiZone').innerHTML        = note;
  document.getElementById('donutsZone').innerHTML     = '';
  document.getElementById('metricsZone').innerHTML    = '';
  document.getElementById('newMemberZone').innerHTML  = '';
  document.getElementById('cancelZone').innerHTML     = '';
}

// ===== MASTER RENDER =====
function render() {
  // Reflect mgmt-filter state on the select (cosmetic — also keeps button-like style in sync)
  const sel = document.getElementById('mgmtSelect');
  if (sel) sel.classList.toggle('filtered', activeMgmt !== ALL_MGMT);

  if (activeMode === 'sxs') {
    // In SXS mode, both panels are independent; only render empty-state if
    // BOTH panels would be empty (i.e. no hotels at all match the mgmt filter).
    const m = HOTELS_ALL.filter(matchesMgmt);
    if (m.length === 0) {
      renderEmptyZones('Try selecting "All Companies" or a different cohort.');
    } else {
      document.getElementById('kpiZone').innerHTML = renderKPISideBySide();
      renderDonutsSXS();
      renderMetricsSXS();
      renderNewMemberSXS();
      renderCancelSXS();
    }
  } else {
    const hotels = activeHotels();
    if (hotels.length === 0) {
      const cohortName = activeMode === 'pilot' ? 'Pilot' : (activeMode === 'nonpilot' ? 'Non-Pilot' : 'Total Portfolio');
      const mgmtName = activeMgmt === ALL_MGMT ? 'any management company' : activeMgmt;
      renderEmptyZones(`The ${cohortName} cohort has no hotels operated by ${mgmtName}. Adjust the cohort or Mgmt Co filter to see results.`);
    } else {
      const agg = aggregateFromHotels(hotels);
      document.getElementById('kpiZone').innerHTML = renderKPISingle(agg);
      renderDonutsSingle(agg);
      renderMetricsSingle(agg);
      renderNewMemberSingle(agg);
      renderCancelSingle(agg);
    }
  }
  renderHotelTable();
  updateTableTitle();
}

// ===== MGMT-CO DROPDOWN INIT =====
function initMgmtDropdown() {
  const sel = document.getElementById('mgmtSelect');
  if (!sel) return;
  sel.innerHTML = MGMT_OPTIONS.map(opt =>
    `<option value="${opt.name.replace(/"/g,'&quot;')}">${opt.name} (${opt.count})</option>`
  ).join('');
  sel.value = ALL_MGMT;
  sel.addEventListener('change', () => {
    activeMgmt = sel.value || ALL_MGMT;
    render();
  });
}

// ===== TOGGLE WIRING =====
document.querySelectorAll('.cohort-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.cohort-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    activeMode = btn.dataset.mode;
    render();
  });
});

// ===== INIT =====
initMgmtDropdown();
render();
</script>
</body>
</html>
"""


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    warnings = []

    print(f"Reading master DB: {MASTER_PATH.name}")
    master = load_master_pilot_map()
    pilot_master_count = sum(1 for v in master.values() if v["pilot"])
    nonpilot_master_count = sum(1 for v in master.values() if not v["pilot"])
    total_master_count = len(master)
    print(f"  master hotels: {total_master_count}  (pilot={pilot_master_count}, non-pilot={nonpilot_master_count})")

    master_stripped = {strip_state_suffix(k): v for k, v in master.items()}

    print(f"Reading rollup:    {ROLLUP_PATH.name}")
    rollup = read_rollup()
    print(f"  rollup keys: {sorted(rollup.keys())}")

    print(f"Reading by-hotel:  {BYHOTEL_PATH.name}")
    records = read_byhotel()
    print(f"  by-hotel records: {len(records)}")

    # Resolve pilot flag + management company per property
    properties = sorted({r["prop"] for r in records})
    pilot_map_by_prop = {}
    mgmt_map_by_prop = {}
    for p in properties:
        pilot_map_by_prop[p] = resolve_pilot(p, master, master_stripped, warnings)
        mgmt_map_by_prop[p]  = resolve_mgmt(p, master, master_stripped, warnings)

    pilot_props    = {p for p, is_pilot in pilot_map_by_prop.items() if is_pilot}
    nonpilot_props = {p for p, is_pilot in pilot_map_by_prop.items() if not is_pilot}
    print(f"  resolved: {len(pilot_props)} pilot, {len(nonpilot_props)} non-pilot, {len(properties)} total")
    # Mgmt-company breakdown across the by-hotel set
    from collections import Counter
    mgmt_counts = Counter(mgmt_map_by_prop.values())
    print(f"  mgmt companies ({len(mgmt_counts)}):")
    for nm, ct in sorted(mgmt_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        print(f"    {nm}: {ct}")

    # Aggregations
    agg_total    = aggregate(records, properties_in_scope=None)
    agg_pilot    = aggregate(records, properties_in_scope=pilot_props)
    agg_nonpilot = aggregate(records, properties_in_scope=nonpilot_props)

    # Compare Total Portfolio against rollup totals
    rollup_pms = rollup.get(("PMS", "Total"), {})
    rollup_web = rollup.get(("WEB", "Total"), {})
    rollup_grand = rollup.get(("Total", "Unknown")) or rollup.get(("Total", "Total"))

    def diff(a, b, key):
        if not b: return 0
        return abs(float(a.get(key, 0)) - float(b.get(key, 0)))

    pms_rev_diff = diff(agg_total["pms"]["total"], rollup_pms, "net_rev")
    web_rev_diff = diff(agg_total["web"]["total"], rollup_web, "net_rev")
    if pms_rev_diff > 1.0:
        warnings.append(f"PMS net_rev mismatch: by-hotel {agg_total['pms']['total']['net_rev']:.2f} vs rollup {rollup_pms.get('net_rev', 0):.2f}")
    if web_rev_diff > 1.0:
        warnings.append(f"WEB net_rev mismatch: by-hotel {agg_total['web']['total']['net_rev']:.2f} vs rollup {rollup_web.get('net_rev', 0):.2f}")
    print(f"  Total Portfolio PMS net_rev vs rollup: ${agg_total['pms']['total']['net_rev']:,.2f} vs ${rollup_pms.get('net_rev', 0):,.2f}  (diff ${pms_rev_diff:.2f})")
    print(f"  Total Portfolio WEB net_rev vs rollup: ${agg_total['web']['total']['net_rev']:,.2f} vs ${rollup_web.get('net_rev', 0):,.2f}  (diff ${web_rev_diff:.2f})")

    # Hotel table rows
    hotel_rows = build_hotel_rows(records, pilot_map_by_prop, mgmt_map_by_prop)

    cohort_counts = {
        "total": total_master_count,
        "pilot": pilot_master_count,
        "nonpilot": nonpilot_master_count,
    }

    cohorts = {
        "total":    _serialize_agg(agg_total),
        "pilot":    _serialize_agg(agg_pilot),
        "nonpilot": _serialize_agg(agg_nonpilot),
    }

    mgmt_options = build_mgmt_options(hotel_rows)

    html = build_html(cohorts, hotel_rows, cohort_counts, mgmt_options, warnings)
    OUTPUT_PATH.write_text(html, encoding="utf-8")
    print(f"\nWrote {OUTPUT_PATH}  ({len(html):,} chars)")

    if warnings:
        print("\nWARNINGS:")
        for w in warnings:
            print(f"  - {w}")

    # Cohort summary print
    print("\n=== Cohort Summary ===")
    for name, a in [("Total Portfolio", agg_total),
                    ("Pilot Cohort",    agg_pilot),
                    ("Non-Pilot Cohort", agg_nonpilot)]:
        g = a["grand"]
        print(f"  {name}: Net Revenue ${g['net_rev']:,.2f} · Net Reservations {g['net_res']:,} · Net ADR ${g['net_adr']:.2f}")

    return 0


def _serialize_agg(a):
    """Round floats for JSON compactness."""
    def r(b):
        out = {}
        for k, v in b.items():
            if isinstance(v, float):
                out[k] = round(v, 4)
            else:
                out[k] = v
        return out
    out = {"pms": {}, "web": {}, "grand": r(a["grand"])}
    for chan in ("pms", "web"):
        for mt, b in a[chan].items():
            out[chan][mt] = r(b)
    return out


if __name__ == "__main__":
    sys.exit(main())
